import psycopg2
import openpyxl
from time import sleep
import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service as ChromeService
from conftest import save_path_screenshot

#import sys
import os
# Agregar el directorio raíz de tu proyecto a la ruta de búsqueda de módulos
#sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from Objetos.objetos_flex import *
from Scripts.funciones import datosJSON, changeFormatDate, getDataExcel, acceptAlert, posIsClosed, selectComboBox, getVoucherNumber

attributes = ["PageFlex","ip","puerto","usuario","contrasenia","baseDeDatos","CasosDePrueba"]
config = datosJSON(r"Escenarios\config.json",attributes)

#Cargo datos de prueba para guardar cierre de cajero
pathExcel = r"DatosDePrueba\guardarCierreCajero.xlsx"
excel = openpyxl.load_workbook(pathExcel)
sheet = excel['Hoja1']

if config["CasosDePrueba"] == []:
    config['CasosDePrueba'] = list(range(2,sheet.max_row + 1))

# Conecto a la base de datos
@pytest.fixture(scope="module")
def cursor():
    db = psycopg2.connect(
    host=config["ip"],
    port=config["puerto"],
    user=config["usuario"],
    password=config["contrasenia"],
    database=config["baseDeDatos"])
    cursor = db.cursor()
    yield cursor
    db.close()

# Configuro driver de selenium
@pytest.fixture(scope="module")
def driver():
    chrome_service = ChromeService("./WebDriver/chromedriver.exe")
    driver = webdriver.Chrome(service=chrome_service)
    driver.maximize_window()

    yield driver
    driver.quit()


# Escenario cierre de cajero
@pytest.mark.parametrize('row',config['CasosDePrueba'])
def test_guardarCierreCajero(driver, cursor, row):
    dataTest = getDataExcel(sheet,row)

    # Constantes de nombres y rutas de carpetas
    folderName = f"{dataTest['Date']} - Pos {dataTest['Pos']} - {dataTest['Shift']}"
    folderPath = f"./ReportesEjecucion/{folderName}/"
    
    # Si la pos no esta cerrada, arroja un error el test
    #posIsClosed(cursor,dataTest["Pos"],dataTest["Date"])

    # Obtengo informacion de la tabla shifts
    cursor.execute(f"""SELECT gross_sale_amount, business_date, pos_number, shift_number FROM shifts WHERE business_date = '{dataTest['Date']}'
                    AND pos_number = {dataTest['Pos']} AND shift_number = {dataTest['Shift']}""")
    register = cursor.fetchone()

    # Verifico haber obtenido informacion
    if register is None:
        pytest.fail("No se encontro el caso de prueba en la base de datos")
    
    # Guardo el monto con decimales correctos
    grossSaleAmount = register[0] / 100

    # Verifico si es "con" o "sin" diferencia
    diffValidate = "con" if dataTest["Diff"] != 0 else "sin" 

    print(f"Cierre de cajero {diffValidate} diferencia")

    # Obtengo la fecha, la pos y el turno por BBDD
    rowDataBase = (str(register[1]),str(register[2]),str(register[3]))
    print(f"Fecha: {register[1]} - Pos: {register[2]} - Turno: {register[3]}")
    
    # Ingreso a la pagina
    driver.get(config['PageFlex'])

    # Si hay una alerta al cambiar de pagina, la acepta para recargar
    acceptAlert(driver)

    # Inicio sesion
    inputBox=input_UserLogin(driver)
    inputBox.send_keys(dataTest["UserFlex"] + Keys.TAB + dataTest["PasswordFlex"] + Keys.ENTER)

    # Ingreso al boton Flex Cash y luego al Cierre de cajero
    btn_FlexCash(driver).click()
    btn_CierreCajero(driver).click()

    # Obtengo los cierre de cajeros pendientes en Flex
    table = table_CierreCajero(driver)
    rows = table.find_elements(By.TAG_NAME,'tr')[1:]
    
    exist=False
    for rowTable in rows:
        columns = rowTable.find_elements(By.TAG_NAME,'td')

        # Obtengo la fecha con formato adecuado, la pos y el turno desde Flex
        rowFlex = (changeFormatDate(columns[0].text),columns[3].text,columns[4].text) 

        # Apreto click cuando consigo la fila correcta
        if rowFlex == rowDataBase:
            rowTable.click()
            exist=True
            break
    if not exist:
        pytest.fail("El ciere de cajero no existe en Flex")

    #Busco el campo para ingresar el dato, borro los 0 e ingreso el monto obtenido
    inputBox = input_PES(driver)
    inputBox.send_keys(Keys.CONTROL + "a" + Keys.BACKSPACE )

    if dataTest["Diff"] is not None:
        grossSaleAmount += dataTest["Diff"]
    
    inputBox.send_keys(f"{grossSaleAmount:.2f}")
    print(f"Se ingresa el siguiente monto en Flex: ${grossSaleAmount:.2f}")
    
    #Si no existe creamos carpeta para guardar las screen separadas por casos de prueba
    if not os.path.exists(folderPath):
        os.makedirs(folderPath)

    #Screen a monto ingresado
    save_path_screenshot(f"{folderName}/monto_ingresado.png")
    driver.save_screenshot(f"{folderPath}monto_ingresado.png")
    

    #Le doy al boton guardar y luego aceptar
    btn_Guardar(driver).click()
    btn_AceptarCierreCajero(driver).click()

    #Guardo screen si hay diferencia
    if dataTest["Diff"] is not None and dataTest["Diff"] !=0:
        btn_AceptarPrePDF(driver).click()
        sleep(2)
        save_path_screenshot(f"{folderName}/monto_diff.png")
        driver.save_screenshot(f"{folderPath}monto_diff.png")

    # Comparo el monto ingresado en Flex con la tabla "transfers", esperando a que la base de datos actualice
    while True:
        sleep(1)
        cursor.execute(f"""
        SELECT amount FROM transfers WHERE origin_pos_number = {dataTest['Pos']} 
        AND business_date = '{dataTest['Date']}' 
        AND origin_shift_number = {dataTest['Shift']}
        AND transfer_type_id = 'POS-A-CIERRE-TURNO' 
        AND tender_name = 'PES'
        """)
        register = cursor.fetchone()
        if register:
            break

    if register[0] / 100 != grossSaleAmount:
        pytest.fail("El valor guardado en Flex no coincide con la base de datos.")
    else:
        print("El valor guardado en Flex coincide con la base de datos.")

    # Ingreso a reimpresión de ticket
    driver.get("http://localhost:6699/tickets/ticketreprinting")
    acceptAlert(driver)
    sleep(2)
    
    # Uso try except porque varían entre 2 o 1 cartel de espera 
    try:
        validateInvisiblyHTML(driver,dataHTML["ticketreprinting"]["wait_Load"],By.XPATH)
    except:
        pass
    sleep(1)
    try:
        validateInvisiblyHTML(driver,dataHTML["ticketreprinting"]["wait_Load2"],By.XPATH)
    except:
        sleep(1)

    # Guardo el boton de busqueda para filtrar
    search = btn_Lupa(driver)

    # Selecciono el filtro de Cierre de cajero y busco
    selectComboBox(comboBox_Operacion(driver),'Cierre de cajero')
    search.click()

    # Obtengo el comprobante utilizando la funcion que filtra la tabla
    nroCierreDeCajero = getVoucherNumber(table_Comprobantes(driver),dataTest)
    
    #Lo guardo en el excel
    sheet.cell(row=int(row), column=7).value = nroCierreDeCajero
    excel.save(pathExcel)
    print(f"Nro. Comprobante Cierre de Cajero: {nroCierreDeCajero}")

    # Si tiene diferencia, entonces tiene amonestacion por lo que obtengo el comprobante correspondiente
    if dataTest["Diff"] is not None and dataTest["Diff"] !=0:
        selectComboBox(comboBox_Operacion(driver),'Amonestación Empleado')
        search.click()
        nroAmonestacion = getVoucherNumber(table_Comprobantes(driver),dataTest)
        sheet.cell(row=int(row), column=8).value = nroAmonestacion
        excel.save(pathExcel)
        print(f"Nro. Comprobante Amonestacion: {nroAmonestacion}")

    return
